import openpyxl
def getRowCount(file, sheetName):
    wb_object = openpyxl.load_workbook(file)
    sheet = wb_object[sheetName]
    return sheet.max_row
def getColumnCount(file, sheetName):
    wb_object = openpyxl.load_workbook(file)
    sheet = wb_object[sheetName]
    return sheet.max_column
def getColumnCount(file, sheetName):
    wb_object = openpyxl.load_workbook(file)
    sheet = wb_object[sheetName]
    return sheet.max_column
def getCellValue(file, sheetName, row_num, col_num):
    wb_object = openpyxl.load_workbook(file)
    sheet = wb_object[sheetName]
    cell_val = sheet.cell(row=row_num, column=col_num).value
    if cell_val == None:
        cell_val = ''
    return (cell_val)

